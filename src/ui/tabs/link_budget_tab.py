"""Link budget tab mixin."""

from __future__ import annotations

import numpy as np
import pyqtgraph as pg
from PySide6.QtCore import Qt
from PySide6.QtGui import QColor, QFont
from PySide6.QtWidgets import (
    QComboBox,
    QDoubleSpinBox,
    QFileDialog,
    QFormLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QMessageBox,
    QPushButton,
    QSplitter,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from src import link_budget_math
from src.itu_losses import estimate_slant_path_loss
from src.models import GroundStationConfig
from src.ui.constants import (
    GIBIT_PER_GBIT,
    LINK_BUDGET_MIN_ELEVATION_DEG,
)


class LinkBudgetTabMixin:
    """Encapsulates link budget UI construction and logic."""

    def _build_link_budget_tab(self) -> QWidget:
        """Assemble the link budget controls, table, and plot."""
        tab = QWidget()
        layout = QHBoxLayout(tab)
        input_group = QGroupBox("Inputs")
        form = QFormLayout(input_group)
        self.link_budget_station_combo = QComboBox()
        form.addRow("Ground station:", self.link_budget_station_combo)
        self.lb_frequency_input = QDoubleSpinBox()
        self.lb_frequency_input.setRange(0.1, 100.0)
        self.lb_frequency_input.setDecimals(3)
        self.lb_frequency_input.setValue(8.2)
        self.lb_frequency_input.setSuffix(" GHz")
        form.addRow("Frequency:", self.lb_frequency_input)
        self.lb_tx_power_input = QDoubleSpinBox()
        self.lb_tx_power_input.setRange(-50.0, 50.0)
        self.lb_tx_power_input.setValue(3.0)
        self.lb_tx_power_input.setSuffix(" dBW")
        form.addRow("TX power:", self.lb_tx_power_input)
        self.lb_tx_gain_input = QDoubleSpinBox()
        self.lb_tx_gain_input.setRange(-10.0, 80.0)
        self.lb_tx_gain_input.setValue(5.41)
        self.lb_tx_gain_input.setSuffix(" dBi")
        form.addRow("TX boresight gain:", self.lb_tx_gain_input)
        self.lb_tx_losses_input = QDoubleSpinBox()
        self.lb_tx_losses_input.setRange(0.0, 20.0)
        self.lb_tx_losses_input.setValue(2.0)
        self.lb_tx_losses_input.setSuffix(" dB")
        form.addRow("TX feeder loss:", self.lb_tx_losses_input)
        self.lb_tx_backoff_input = QDoubleSpinBox()
        self.lb_tx_backoff_input.setRange(0.0, 10.0)
        self.lb_tx_backoff_input.setValue(0.0)
        self.lb_tx_backoff_input.setSuffix(" dB")
        form.addRow("TX backoff:", self.lb_tx_backoff_input)
        self.lb_antenna_gain_input = QDoubleSpinBox()
        self.lb_antenna_gain_input.setRange(-10.0, 80.0)
        self.lb_antenna_gain_input.setValue(5.41)
        self.lb_antenna_gain_input.setSuffix(" dBi")
        form.addRow("Actual antenna gain:", self.lb_antenna_gain_input)
        self.lb_rx_gt_input = QDoubleSpinBox()
        self.lb_rx_gt_input.setRange(-10.0, 80.0)
        self.lb_rx_gt_input.setValue(26.0)
        self.lb_rx_gt_input.setSuffix(" dB/K")
        form.addRow("Receiver G/T:", self.lb_rx_gt_input)
        self.lb_rx_losses_input = QDoubleSpinBox()
        self.lb_rx_losses_input.setRange(0.0, 20.0)
        self.lb_rx_losses_input.setValue(0.0)
        self.lb_rx_losses_input.setSuffix(" dB")
        form.addRow("Receiver losses:", self.lb_rx_losses_input)
        self.lb_symbol_rate_input = QDoubleSpinBox()
        self.lb_symbol_rate_input.setRange(0.01, 5000.0)
        self.lb_symbol_rate_input.setValue(300.0)
        self.lb_symbol_rate_input.setDecimals(3)
        self.lb_symbol_rate_input.setSuffix(" Msps")
        form.addRow("Symbol rate limit:", self.lb_symbol_rate_input)
        self.lb_impl_loss_input = QDoubleSpinBox()
        self.lb_impl_loss_input.setRange(0.0, 10.0)
        self.lb_impl_loss_input.setValue(1.0)
        self.lb_impl_loss_input.setSuffix(" dB")
        form.addRow("Implementation loss:", self.lb_impl_loss_input)
        self.lb_margin_input = QDoubleSpinBox()
        self.lb_margin_input.setRange(0.0, 20.0)
        self.lb_margin_input.setValue(3.0)
        self.lb_margin_input.setSuffix(" dB")
        form.addRow("Link margin target:", self.lb_margin_input)
        self.lb_sat_altitude_input = QDoubleSpinBox()
        self.lb_sat_altitude_input.setRange(150.0, 50000.0)
        self.lb_sat_altitude_input.setValue(550.0)
        self.lb_sat_altitude_input.setSuffix(" km")
        form.addRow("Satellite altitude:", self.lb_sat_altitude_input)
        self.lb_gs_elevation_input = QDoubleSpinBox()
        self.lb_gs_elevation_input.setRange(0.0, 90.0)
        self.lb_gs_elevation_input.setValue(60.0)
        self.lb_gs_elevation_input.setSuffix(" °")
        form.addRow("GS elevation angle:", self.lb_gs_elevation_input)
        self.lb_unavailability_input = QDoubleSpinBox()
        self.lb_unavailability_input.setRange(0.01, 5.0)
        self.lb_unavailability_input.setValue(0.1)
        self.lb_unavailability_input.setDecimals(2)
        self.lb_unavailability_input.setSuffix(" %")
        form.addRow("Unavailability:", self.lb_unavailability_input)
        self.lb_polarization_loss_input = QDoubleSpinBox()
        self.lb_polarization_loss_input.setRange(0.0, 5.0)
        self.lb_polarization_loss_input.setDecimals(2)
        self.lb_polarization_loss_input.setValue(0.1)
        self.lb_polarization_loss_input.setSuffix(" dB")
        form.addRow("Polarization loss:", self.lb_polarization_loss_input)
        self.lb_rolloff_input = QDoubleSpinBox()
        self.lb_rolloff_input.setRange(0.05, 1.0)
        self.lb_rolloff_input.setDecimals(2)
        self.lb_rolloff_input.setValue(0.25)
        form.addRow("Roll-off factor:", self.lb_rolloff_input)
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_layout.addWidget(input_group)
        left_layout.addWidget(self._build_downlink_summary_group())
        left_layout.addStretch(1)
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        self.link_budget_tabs = QTabWidget()
        right_layout.addWidget(self.link_budget_tabs, stretch=1)
        table_tab = QWidget()
        table_layout = QVBoxLayout(table_tab)
        self.link_budget_table = QTableWidget(0, 3)
        self.link_budget_table.setHorizontalHeaderLabels(["Parameter", "Value", "Unit"])
        self.link_budget_table.verticalHeader().setVisible(False)
        self.link_budget_table.setAlternatingRowColors(True)
        self.link_budget_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.link_budget_table.setSelectionMode(QTableWidget.SelectionMode.NoSelection)
        self.link_budget_table.setWordWrap(False)
        header = self.link_budget_table.horizontalHeader()
        header.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        table_layout.addWidget(self.link_budget_table)
        self.link_budget_summary_label = QLabel(
            "Adjust the inputs to compute the link budget."
        )
        self.link_budget_summary_label.setWordWrap(True)
        table_layout.addWidget(self.link_budget_summary_label)
        export_button = QPushButton("Export to Excel")
        export_button.clicked.connect(self._export_link_budget_to_xlsx)
        table_layout.addWidget(export_button)
        plot_tab = QWidget()
        plot_layout = QVBoxLayout(plot_tab)
        self.link_budget_plot = pg.PlotWidget(title="VCM System Performance")
        self.link_budget_plot.setLabel("bottom", "Elevation", units="deg")
        self.link_budget_plot.setLabel("left", "Es/N0", units="dB")
        self.link_budget_plot.showGrid(x=True, y=True, alpha=0.3)
        self._link_budget_plot_legend = self.link_budget_plot.addLegend(offset=(10, 10))
        self._link_budget_plot_annotations: list = []
        plot_layout.addWidget(self.link_budget_plot)
        self.link_budget_tabs.addTab(table_tab, "Static Link Budget")
        self.link_budget_tabs.addTab(plot_tab, "Dynamic Link Budget")
        self._dynamic_link_budget_tab = plot_tab
        self._latest_link_budget_plot_data: tuple[np.ndarray, dict] | None = None
        self._loss_cache_key: tuple | None = None
        self._loss_cache_losses: np.ndarray | None = None
        self._loss_cache_contributions: dict | None = None
        self.link_budget_tabs.currentChanged.connect(self._on_link_budget_tab_changed)
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.addWidget(left_widget)
        splitter.addWidget(right_widget)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 1)
        splitter.setSizes([1, 1])
        layout.addWidget(splitter)
        self._register_link_budget_inputs()
        self._refresh_link_budget_station_list()
        self._link_budget_auto_enabled = True
        self._trigger_link_budget_recompute()
        return tab

    def _build_downlink_summary_group(self) -> QGroupBox:
        """Create the box summarizing downlink capacity."""
        group = QGroupBox("Downlink Summary")
        layout = QFormLayout(group)
        self._downlink_total_label = QLabel("—")
        self._downlink_per_orbit_label = QLabel("—")
        layout.addRow("Scenario total:", self._downlink_total_label)
        layout.addRow("Per orbit:", self._downlink_per_orbit_label)
        return group

    def _get_selected_link_budget_station(self) -> GroundStationConfig | None:
        """Return the station referenced by the link-budget dropdown."""
        combo = self.link_budget_station_combo
        if combo is None or combo.count() == 0:
            return None
        data = combo.currentData()
        if isinstance(data, int) and 0 <= data < len(self._station_presets):
            return self._station_presets[data]
        return None

    def _refresh_link_budget_station_list(self) -> None:
        """Keep the link-budget dropdown in sync with the station list."""
        combo = self.link_budget_station_combo
        if combo is None:
            return
        current_data = combo.currentData()
        combo.blockSignals(True)
        combo.clear()
        if not self._station_presets:
            combo.addItem("No stations loaded", None)
            combo.setEnabled(False)
        else:
            for idx, station in enumerate(self._station_presets):
                combo.addItem(station.name, idx)
            combo.setEnabled(True)
            if isinstance(current_data, int) and 0 <= current_data < len(
                self._station_presets
            ):
                combo.setCurrentIndex(current_data)
        combo.blockSignals(False)
        self._trigger_link_budget_recompute()

    def _register_link_budget_inputs(self) -> None:
        """Connect all link-budget controls to automatic recomputation."""
        controls = [
            (self.link_budget_station_combo, "currentIndexChanged"),
            (self.lb_frequency_input, "valueChanged"),
            (self.lb_tx_power_input, "valueChanged"),
            (self.lb_tx_gain_input, "valueChanged"),
            (self.lb_tx_losses_input, "valueChanged"),
            (self.lb_tx_backoff_input, "valueChanged"),
            (self.lb_antenna_gain_input, "valueChanged"),
            (self.lb_rx_gt_input, "valueChanged"),
            (self.lb_rx_losses_input, "valueChanged"),
            (self.lb_symbol_rate_input, "valueChanged"),
            (self.lb_impl_loss_input, "valueChanged"),
            (self.lb_margin_input, "valueChanged"),
            (self.lb_sat_altitude_input, "valueChanged"),
            (self.lb_gs_elevation_input, "valueChanged"),
            (self.lb_unavailability_input, "valueChanged"),
            (self.lb_polarization_loss_input, "valueChanged"),
            (self.lb_rolloff_input, "valueChanged"),
        ]
        for widget, signal_name in controls:
            self._connect_link_budget_signal(widget, signal_name)

    def _connect_link_budget_signal(self, widget, signal_name: str) -> None:
        """Attach the specified signal to trigger recomputation."""
        if widget is None:
            return
        signal = getattr(widget, signal_name, None)
        if signal is None:
            return
        signal.connect(self._trigger_link_budget_recompute)  # type: ignore[attr-defined]

    def _trigger_link_budget_recompute(self, *_args) -> None:
        """Recalculate the link budget when inputs change."""
        if not self._link_budget_auto_enabled:
            return
        self._handle_link_budget_calculate()

    def _handle_link_budget_calculate(self) -> None:
        """Compute the link budget using the math helpers."""
        if (
            self.link_budget_table is None
            or self.link_budget_summary_label is None
            or self.lb_frequency_input is None
        ):
            return
        station = self._get_selected_link_budget_station()
        if station is None:
            if self.link_budget_summary_label:
                self.link_budget_summary_label.setText(
                    "Import or select a ground station to view the link budget."
                )
            if self.link_budget_table:
                self.link_budget_table.setRowCount(0)
            self._clear_link_budget_plot()
            self._link_budget_rate_curve = None
            self._update_downlink_summary()
            self._invalidate_loss_cache()
            self._latest_link_budget_plot_data = None
            return
        frequency = self.lb_frequency_input.value()
        symbol_rate_sps = self.lb_symbol_rate_input.value() * 1e6
        min_elevation = LINK_BUDGET_MIN_ELEVATION_DEG
        plot_lower_bound = LINK_BUDGET_MIN_ELEVATION_DEG
        evaluation_angle = self.lb_gs_elevation_input.value()
        evaluation_angle = max(plot_lower_bound, min(90.0, evaluation_angle))
        unavailability = self.lb_unavailability_input.value()
        try:
            elevations = np.linspace(plot_lower_bound, 90.0, 1000)
            loss_cache_key = self._build_loss_cache_key(
                frequency=frequency,
                plot_lower_bound=plot_lower_bound,
                unavailability=unavailability,
                station=station,
                num_samples=elevations.size,
            )
            use_cache = (
                self._loss_cache_key == loss_cache_key
                and self._loss_cache_losses is not None
                and self._loss_cache_losses.shape == elevations.shape
            )
            if use_cache:
                atmospheric_losses = self._loss_cache_losses
                contribution_breakdown = self._loss_cache_contributions
            else:
                loss_result = estimate_slant_path_loss(
                    frequency_GHz=frequency,
                    elevations_deg=elevations,
                    lat_deg=station.latitude_deg,
                    lon_deg=station.longitude_deg,
                    altitude_m=station.altitude_m,
                    unavailability_percent=unavailability,
                    return_contributions=True,
                )
                if isinstance(loss_result, tuple):
                    atmospheric_losses, contribution_breakdown = loss_result
                else:
                    atmospheric_losses = loss_result
                    contribution_breakdown = None
                self._loss_cache_key = loss_cache_key
                self._loss_cache_losses = atmospheric_losses
                self._loss_cache_contributions = contribution_breakdown
            results = link_budget_math.calculate_link_budget(
                elevations_deg=elevations,
                antenna_gains_dBi=np.full_like(
                    elevations, self.lb_antenna_gain_input.value()
                ),
                atmospheric_losses_dB=atmospheric_losses,
                tx_power_dBw=self.lb_tx_power_input.value(),
                tx_boresight_gain_dBi=self.lb_tx_gain_input.value(),
                tx_losses_dB=self.lb_tx_losses_input.value(),
                tx_backoff_dB=self.lb_tx_backoff_input.value(),
                frequency_GHz=frequency,
                satellite_altitude_km=self.lb_sat_altitude_input.value(),
                ground_altitude_m=station.altitude_m,
                receiver_G_T_dB_K=self.lb_rx_gt_input.value(),
                receiver_losses_dB=self.lb_rx_losses_input.value(),
                symbol_rate_sps=symbol_rate_sps,
                implementation_loss_dB=self.lb_impl_loss_input.value(),
                margin_dB=self.lb_margin_input.value(),
            )
            rows = link_budget_math.build_parameter_rows(
                elevations_deg=elevations,
                results=results,
                evaluation_elevation_deg=evaluation_angle,
                min_gs_elevation_deg=min_elevation,
                tx_frequency_GHz=frequency,
                tx_power_dBw=self.lb_tx_power_input.value(),
                tx_losses_dB=self.lb_tx_losses_input.value(),
                tx_boresight_gain_dBi=self.lb_tx_gain_input.value(),
                tx_backoff_dB=self.lb_tx_backoff_input.value(),
                symbol_rate_sps=symbol_rate_sps,
                receiver_G_T_dB_K=self.lb_rx_gt_input.value(),
                implementation_loss_dB=self.lb_impl_loss_input.value(),
                margin_dB=self.lb_margin_input.value(),
                rolloff=self.lb_rolloff_input.value(),
                polarization_loss_dB=self.lb_polarization_loss_input.value(),
                satellite_altitude_km=self.lb_sat_altitude_input.value(),
                atmospheric_breakdown_dB=contribution_breakdown,
            )
        except Exception as exc:
            QMessageBox.warning(self, "Link Budget Error", str(exc))
            self._clear_link_budget_plot()
            self._latest_link_budget_plot_data = None
            self._link_budget_rate_curve = None
            self._update_downlink_summary()
            return
        eval_index = int(np.argmin(np.abs(elevations - evaluation_angle)))
        summary = self._format_link_budget_summary(results, elevations, eval_index)
        self._populate_link_budget_table(rows, summary)
        self._latest_link_budget_plot_data = (elevations, results)
        if self._is_dynamic_link_budget_tab_active():
            self._update_link_budget_plot(elevations, results)
        self._cache_link_budget_curve(elevations, results)
        self._update_downlink_summary()

    def _populate_link_budget_table(
        self,
        rows: list[link_budget_math.ParameterRow],
        summary_text: str,
    ) -> None:
        """Fill the link-budget table and summary label."""
        if not self.link_budget_table or self.link_budget_summary_label is None:
            return
        table = self.link_budget_table
        table.setRowCount(len(rows))
        highlight_rows = {
            "EIRP",
            "Received Signal Power",
            "Maximum Information Rate",
            "Max. Information Rate",
            "Link Margin",
        }
        highlight_color = QColor("#1b5e20")
        for row_idx, entry in enumerate(rows):
            items = [
                QTableWidgetItem(entry.parameter),
                QTableWidgetItem(entry.value),
                QTableWidgetItem(entry.unit),
            ]
            for col, item in enumerate(items):
                if col == 0:
                    align = Qt.AlignmentFlag.AlignLeft
                elif col == 1:
                    align = Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
                else:
                    align = Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter
                item.setTextAlignment(align)
                table.setItem(row_idx, col, item)
            if entry.parameter in highlight_rows:
                for item in items:
                    font = QFont(item.font())
                    font.setBold(True)
                    item.setFont(font)
                    item.setBackground(highlight_color)
        if not rows:
            self.link_budget_summary_label.setText(
                "No valid MODCOD found for the current parameters."
            )
        else:
            self.link_budget_summary_label.setText(summary_text)

    def _format_link_budget_summary(
        self,
        results: dict,
        elevations: np.ndarray,
        index: int,
    ) -> str:
        """Generate a concise summary for the selected elevation."""
        elevation_deg = float(elevations[index])
        modcod = results["modcod_names"][index]
        data_rate = float(np.asarray(results["data_rate_mbps"])[index])
        margin = float(np.asarray(results["margin_to_required_EsN0_dB"])[index])
        if modcod == "No Link" or np.isnan(margin):
            return (
                f"No MODCOD closes at {elevation_deg:.1f}°. "
                f"Available data rate: {data_rate:.2f} Mbps."
            )
        return (
            f"{modcod} at {elevation_deg:.1f}° delivers {data_rate:.2f} Mbps "
            f"(margin {margin:.2f} dB to threshold)."
        )

    def _clear_link_budget_plot(self) -> None:
        """Remove plot data and annotations."""
        if not getattr(self, "link_budget_plot", None):
            return
        self.link_budget_plot.clear()
        if getattr(self, "_link_budget_plot_legend", None):
            self._link_budget_plot_legend.clear()
        for item in getattr(self, "_link_budget_plot_annotations", []):
            try:
                self.link_budget_plot.removeItem(item)
            except Exception:
                pass
        self._link_budget_plot_annotations = []

    def _invalidate_loss_cache(self) -> None:
        self._loss_cache_key = None
        self._loss_cache_losses = None
        self._loss_cache_contributions = None

    def _build_loss_cache_key(
        self,
        *,
        frequency: float,
        plot_lower_bound: float,
        unavailability: float,
        station: GroundStationConfig,
        num_samples: int,
    ) -> tuple:
        return (
            round(float(frequency), 6),
            round(float(plot_lower_bound), 4),
            round(float(unavailability), 6),
            round(float(station.latitude_deg), 6),
            round(float(station.longitude_deg), 6),
            round(float(station.altitude_m), 3),
            int(num_samples),
        )

    def _is_dynamic_link_budget_tab_active(self) -> bool:
        return (
            self.link_budget_tabs is not None
            and getattr(self, "_dynamic_link_budget_tab", None) is not None
            and self.link_budget_tabs.currentWidget() is self._dynamic_link_budget_tab
        )

    def _on_link_budget_tab_changed(self, index: int) -> None:
        if not self._is_dynamic_link_budget_tab_active():
            return
        if not self._latest_link_budget_plot_data:
            return
        elevations, results = self._latest_link_budget_plot_data
        self._update_link_budget_plot(elevations, results)

    def _update_link_budget_plot(
        self,
        elevations: np.ndarray,
        results: dict,
    ) -> None:
        """Render the Es/N0 envelope as a function of elevation."""
        if not getattr(self, "link_budget_plot", None):
            return
        if elevations.size == 0:
            self._clear_link_budget_plot()
            return
        es_n0 = np.asarray(results.get("es_to_n0_dB", []), dtype=float)
        required = np.asarray(results.get("required_EsN0_dB", []), dtype=float)
        if es_n0.size != elevations.size:
            self._clear_link_budget_plot()
            return
        margin_offset = self.lb_margin_input.value()

        def _build_step_curve(
            x_values: np.ndarray, y_values: np.ndarray
        ) -> tuple[np.ndarray, np.ndarray]:
            if len(x_values) == 0 or len(y_values) == 0:
                return x_values, y_values
            repeated_x = np.repeat(x_values, 2)[1:]
            repeated_y = np.repeat(y_values, 2)[:-1]
            return repeated_x, repeated_y

        self.link_budget_plot.clear()
        if getattr(self, "_link_budget_plot_legend", None):
            self._link_budget_plot_legend.clear()
        for item in self._link_budget_plot_annotations:
            try:
                self.link_budget_plot.removeItem(item)
            except Exception:
                pass
        self._link_budget_plot_annotations = []
        self.link_budget_plot.plot(
            elevations,
            es_n0,
            pen=pg.mkPen("#ff9800", width=2),
            name="Maximum Es/N0",
        )
        modcods = results.get("modcod_names", [])
        valid_throughput = np.where(
            ~np.isnan(required), required + margin_offset, np.nan
        )
        if len(modcods) == len(valid_throughput):
            valid_mask = np.array([name != "No Link" for name in modcods], dtype=bool)
            valid_throughput = np.where(valid_mask, valid_throughput, np.nan)
        step_x, step_y = _build_step_curve(elevations, valid_throughput)
        self.link_budget_plot.plot(
            step_x,
            step_y,
            pen=pg.mkPen("#5b7dff", width=2),
            name="VCM Step",
        )
        valid_margin = es_n0 - valid_throughput
        self.link_budget_plot.plot(
            elevations,
            valid_margin,
            pen=pg.mkPen("#00c853", width=2),
            name="VCM Margin",
        )
        y_min = -5.0
        y_max = 20.0
        self.link_budget_plot.setYRange(y_min, y_max, padding=0.0)
        self.link_budget_plot.setLimits(yMin=y_min, yMax=y_max)
        modcods = results.get("modcod_names", [])
        if len(modcods) == len(elevations):
            start_idx = 0
            current_name = modcods[0]
            segments: list[tuple[int, int, str]] = []
            for idx in range(1, len(modcods)):
                if modcods[idx] != current_name:
                    segments.append((start_idx, idx - 1, current_name))
                    start_idx = idx
                    current_name = modcods[idx]
            segments.append((start_idx, len(modcods) - 1, current_name))
            for start, end, name in segments:
                if name == "No Link":
                    continue
                segment_values = valid_throughput[start : end + 1]
                mean_value = np.nanmean(segment_values)
                if np.isnan(mean_value):
                    continue
                center_x = (elevations[start] + elevations[end]) / 2.0
                text_item = pg.TextItem(name, color="#d0d0d0", anchor=(0.5, -0.3))
                text_item.setPos(center_x, mean_value)
                self.link_budget_plot.addItem(text_item)
                self._link_budget_plot_annotations.append(text_item)
        if len(modcods) == len(elevations):
            modcod_array = np.asarray(modcods, dtype=object)
            linked_indices = np.where(modcod_array != "No Link")[0]
            if linked_indices.size > 0:
                boundary_idx = int(linked_indices[0])
                if boundary_idx > 0 and np.all(
                    modcod_array[:boundary_idx] == "No Link"
                ):
                    boundary_x = float(elevations[boundary_idx])
                    shade = pg.LinearRegionItem(
                        values=(float(elevations[0]), boundary_x),
                        brush=pg.mkBrush(255, 77, 77, 60),
                        movable=False,
                    )
                    shade.setZValue(-100)
                    self.link_budget_plot.addItem(shade)
                    self._link_budget_plot_annotations.append(shade)
                    line_pen = pg.mkPen("#ff4d4d", width=2, style=Qt.PenStyle.DotLine)
                    boundary_line = pg.InfiniteLine(
                        pos=boundary_x, angle=90, pen=line_pen
                    )
                    self.link_budget_plot.addItem(boundary_line)
                    self._link_budget_plot_annotations.append(boundary_line)

    def _store_access_series(self, result) -> None:
        """Cache the elevation time series for downlink calculations."""
        if not result.timeline_seconds or not result.station_elevation_series:
            self._latest_access_series = None
            return
        time_seconds = np.asarray(result.timeline_seconds, dtype=float)
        if time_seconds.ndim != 1 or time_seconds.size == 0:
            self._latest_access_series = None
            return
        station_series: dict[str, np.ndarray] = {}
        for name, samples in result.station_elevation_series.items():
            series = np.asarray(samples, dtype=float)
            if series.size != time_seconds.size:
                continue
            station_series[name] = series
        if not station_series:
            self._latest_access_series = None
            return
        self._latest_access_series = {
            "time_seconds": time_seconds,
            "station_series": station_series,
            "orbit_period_s": float(result.orbit_period_seconds or 0.0),
        }

    def _cache_link_budget_curve(self, elevations: np.ndarray, results: dict) -> None:
        """Store the latest throughput curve for interpolation."""
        rates = np.asarray(results.get("data_rate_mbps", []), dtype=float)
        if elevations.size and rates.size == elevations.size:
            self._link_budget_rate_curve = (np.asarray(elevations, dtype=float), rates)
        else:
            self._link_budget_rate_curve = None

    def _update_downlink_summary(self) -> None:
        """Refresh the downlink summary labels."""
        if not self._downlink_total_label or not self._downlink_per_orbit_label:
            return
        metrics = self._compute_downlink_metrics()
        if metrics is None:
            self._downlink_total_label.setText("—")
            self._downlink_per_orbit_label.setText("—")
            return
        total_gibit, per_orbit_gibit = metrics
        self._downlink_total_label.setText(f"{total_gibit:.2f} Gibit")
        if per_orbit_gibit is None:
            self._downlink_per_orbit_label.setText("—")
        else:
            self._downlink_per_orbit_label.setText(f"{per_orbit_gibit:.2f} Gibit/orbit")

    def _compute_downlink_metrics(self) -> tuple[float, float | None] | None:
        """Compute total and per-orbit Gibit using cached data."""
        if not self._latest_access_series or not self._link_budget_rate_curve:
            return None
        time_seconds = np.asarray(
            self._latest_access_series.get("time_seconds", []), dtype=float
        )
        station_series: dict[str, np.ndarray] = self._latest_access_series.get(
            "station_series", {}
        )
        orbit_period_s = float(self._latest_access_series.get("orbit_period_s", 0.0))
        if time_seconds.size < 2 or not station_series:
            return None
        grid, rates = self._link_budget_rate_curve
        if grid.size == 0 or rates.size == 0:
            return None
        total_gbit = 0.0
        for series in station_series.values():
            if series.size != time_seconds.size:
                continue
            station_rates = np.interp(
                np.nan_to_num(series, nan=0.0),
                grid,
                rates,
                left=0.0,
                right=0.0,
            )
            station_rates = np.nan_to_num(station_rates, nan=0.0)
            if not np.any(station_rates):
                continue
            gbit = link_budget_math.integrate_data_volume_gb(
                time_seconds, station_rates
            )
            total_gbit += max(0.0, float(gbit))
        if total_gbit <= 0.0:
            return (0.0, None)
        total_gibit = total_gbit * GIBIT_PER_GBIT
        duration_s = float(time_seconds[-1] - time_seconds[0])
        per_orbit = None
        if orbit_period_s > 0.0:
            num_orbits = duration_s / orbit_period_s
            if num_orbits > 0:
                per_orbit = total_gibit / num_orbits
        return (total_gibit, per_orbit)

    def _export_link_budget_to_xlsx(self) -> None:
        """Export the link budget table to an Excel file with formatting."""
        if not self.link_budget_table or self.link_budget_table.rowCount() == 0:
            QMessageBox.warning(
                self,
                "Export Error",
                "No link budget data to export. Please calculate the link budget first.",
            )
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            QMessageBox.critical(
                self,
                "Import Error",
                "openpyxl is not installed. Please install it using: pip install openpyxl",
            )
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Link Budget",
            "link_budget.xlsx",
            "Excel Files (*.xlsx)",
        )

        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Link Budget"

            highlight_rows = {
                "EIRP",
                "Received Signal Power",
                "Maximum Information Rate",
                "Max. Information Rate",
                "Link Margin",
            }
            highlight_color = "1B5E20"

            headers = []
            for col in range(self.link_budget_table.columnCount()):
                header_item = self.link_budget_table.horizontalHeaderItem(col)
                headers.append(header_item.text() if header_item else "")
            ws.append(headers)

            for col_idx, _ in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for row in range(self.link_budget_table.rowCount()):
                row_data = []
                for col in range(self.link_budget_table.columnCount()):
                    item = self.link_budget_table.item(row, col)
                    row_data.append(item.text() if item else "")
                ws.append(row_data)

                parameter_name = row_data[0] if row_data else ""
                is_highlight = parameter_name in highlight_rows

                for col_idx in range(1, self.link_budget_table.columnCount() + 1):
                    cell = ws.cell(row=row + 2, column=col_idx)

                    if col_idx == 1:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    elif col_idx == 2:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

                    if is_highlight:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(
                            start_color=highlight_color,
                            end_color=highlight_color,
                            fill_type="solid",
                        )

            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 15
            ws.column_dimensions["C"].width = 12

            wb.save(file_path)
            QMessageBox.information(
                self,
                "Export Successful",
                f"Link budget exported successfully to:\n{file_path}",
            )

        except Exception as exc:
            QMessageBox.critical(
                self,
                "Export Error",
                f"Failed to export link budget:\n{str(exc)}",
            )

